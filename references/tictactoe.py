# srp tic tac toe sim
# ties = 16/958
import random
import itertools
from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
import csv
# import tensorflow

endgames = []

# filename = openpyxl.load_workbook("Tic_Tac_Toe.xlsx")
# sheetname = filename.sheetnames
wb = Workbook()
ws = wb.active
# filename = "Tic_Tac_Toe.csv"
headers = ['Score', 'Speed', 'x', 'o', 'empty']

while len(endgames) < 958:

    blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
    choice_x = []
    choice_o = []
    finish_x = False
    finish_o = False
    matches = []

    for x, y in itertools.product(blocks, blocks):

        if len(blocks) >= 2:

            chosen_block = blocks[random.randint(0, len(blocks) - 1)]
            choice_x.append(chosen_block)
            blocks.pop(blocks.index(chosen_block))
            # if detection under here
            if 'a' in choice_x and 'b' in choice_x and 'c' in choice_x:
                finish_x = True
            if 'd' in choice_x and 'e' in choice_x and 'f' in choice_x:
                finish_x = True
            if 'g' in choice_x and 'h' in choice_x and 'i' in choice_x:
                finish_x = True
            if 'a' in choice_x and 'd' in choice_x and 'g' in choice_x:
                finish_x = True
            if 'c' in choice_x and 'e' in choice_x and 'g' in choice_x:
                finish_x = True
            if 'b' in choice_x and 'e' in choice_x and 'h' in choice_x:
                finish_x = True
            if 'c' in choice_x and 'f' in choice_x and 'i' in choice_x:
                finish_x = True
            if 'a' in choice_x and 'e' in choice_x and 'i' in choice_x:
                finish_x = True
            if finish_x:
                matches.append(1)
                break

            chosen_block = blocks[random.randint(0, len(blocks) - 1)]
            choice_o.append(chosen_block)
            blocks.pop(blocks.index(chosen_block))
            # if detection under here
            if 'a' in choice_o and 'b' in choice_o:
                if 'c' in choice_o:
                    finish_o = True
            if 'd' in choice_o and 'e' in choice_o:
                if 'f' in choice_o:
                    finish_o = True
            if 'g' in choice_o and 'h' in choice_o:
                if 'i' in choice_o:
                    finish_o = True
            if 'a' in choice_o and 'd' in choice_o and 'g' in choice_o:
                finish_o = True
            if 'c' in choice_o and 'e' in choice_o and 'g' in choice_o:
                finish_o = True
            if 'b' in choice_o and 'e' in choice_o:
                if 'h' in choice_o:
                    finish_o = True
            if 'c' in choice_o and 'f' in choice_o and 'i' in choice_o:
                finish_o = True
            if 'a' in choice_o and 'e' in choice_o and 'i' in choice_o:
                finish_o = True
            if finish_o:
                matches.append(-1)
                break

        if len(blocks) < 2:
            choice_x.append(blocks[0])
            blocks.pop()
            if 'a' in choice_x:
                if 'b' in choice_x:
                    if 'c' in choice_x:
                        finish_x = True
            if 'd' in choice_x and 'e' in choice_x:
                if 'f' in choice_x:
                    finish_x = True
            if 'g' in choice_x and 'h' in choice_x:
                if 'i' in choice_x:
                    finish_x = True
            if 'a' in choice_x and 'd' in choice_x:
                if 'g' in choice_x:
                    finish_x = True

            if 'c' in choice_x and 'e' in choice_x and 'g' in choice_x:
                finish_x = True
            if 'b' in choice_x and 'e' in choice_x and 'h' in choice_x:
                finish_x = True
            if 'c' in choice_x and 'f' in choice_x and 'i' in choice_x:
                finish_x = True
            if 'a' in choice_x and 'e' in choice_x and 'i' in choice_x:
                finish_x = True
            if finish_x:
                matches.append(1)
                break
            else:
                matches.append(0)
                break

    choice_x.sort()
    choice_o.sort()

    matches.append(len(blocks))
    matches.append(choice_x)
    matches.append(choice_o)
    matches.append(blocks)
    # matches = str(matches)
    # matches = matches[1:-1]
    # # matches = '('+matches+')'

    if matches in endgames:
        matches = []
    else:
        endgames.append(matches)
    matches = []

tie_count = 0
t2total_ratio = []
for match in endgames:
    if match[0] == 0:
        tie_count += 1
        t2total_ratio.append(tie_count/(endgames.index(match)+1))
print(t2total_ratio)

# endgames.sort(reverse=True)

# for match in endgames:
#     print(match)
# print(len(endgames))


# def writer(headers, endgames, filename):
#     with open(filename, 'w', newline="") as csvfile:
#         match = csv.writer(csvfile)
#         match.writerow(headers)
#         for alpha in endgames:
#             match.writerow(alpha)
    # y = ['A', 'B', 'C', 'D', 'E']
    # for alpha in range(1, 960):
    #     for beta in range(0, 5):
    #         if alpha == 1:
    #             ws[f'{y[beta]}{alpha}'] = headers[beta]
    #         else:
    #             ws[f'{y[beta]}{alpha}'] = str(endgames[alpha-2][beta])
    #
    # wb.save('dataset.xlsx')
# maybe put the slower ones for losing higher later on
# take the string out, take away first and last char which are the brackets, then .join after split by ,
# create a list with all the winning combos, or maybe a loop to test and see if the output has all that it needs
