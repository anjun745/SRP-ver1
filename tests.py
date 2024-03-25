# import matplotlib
# from openpyxl import Workbook
# import matplotlib.pyplot as plt
# from openpyxl.utils import get_column_letter
# import numpy as np
# import random
# import datetime
from numba import jit
import xlrd
import time
from openpyxl import Workbook

x_blocks = {'x': []}
o_blocks = {'o': []}
blocks = ['c', 'd', 'e', 'g', 'h', 'i']
location1 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\speedtest0.xlsx'
location2 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\speedtest1.xlsx'
location3 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\speedtest2.xlsx'
# location1 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\collection_1.xlsx'
# location2 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\collection_1.xlsx'
# location3 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\collection_1.xlsx'
wins = (['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'],
        ['a', 'd', 'g'], ['b', 'e', 'h'], ['c', 'f', 'i'],
        ['a', 'e', 'i'], ['c', 'e', 'g'])


def win_checker(placed, plc=None):
    if placed == x_blocks:
        placed = x_blocks['x']
    else:
        placed = o_blocks['o']
    for win in wins:
        counter = 0
        for block in placed:
            if block in win:
                counter += 1
        if counter == 3:
            return 'win'
        if plc and counter == 2:
            return win
    return 'ongoing'


# @jit(nopython=True)
def reducing(loc=None, lvl2=None):
    # if not loc:
    #     loc = input('path: ')
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    winning_x = []
    winning_o = []
    draw_counts = []
    lists_of_plays = [winning_x, winning_o, draw_counts]
    rows_to_do = [1, 2, 3]
    for r in rows_to_do:
        if not sheet.cell_value(2, r):
            continue
        for combo in range(int(sheet.cell_value(r, 4)) - 1):
            com = sheet.cell_value(combo + 1, r)
            if lvl2:
                com = com.replace('{', '').replace('}', '').replace("'x': ", '').replace("'o': ", '')
            com = com[2:-2].replace("'", '').replace(', ', '').split('][')
            # comb = [[list(com[0])], [list(com[1])]]
            comb = [[], []]
            for block in com[0]:
                comb[0].append(block)
            for block in com[1]:
                comb[1].append(block)
            if comb not in lists_of_plays[rows_to_do.index(r)]:  # do a for loop checking if there are ones dif order
                lists_of_plays[rows_to_do.index(r)].append(comb)
    return lists_of_plays


# @jit(nopython=True)
def combining(list1, list2, list3):
    combined = []
    lists = [list1, list2, list3]
    indexes = (0, 1, 2)
    for ind in indexes:
        temp_list = []
        for list_ in lists:
            for combo in list_[ind]:
                if combo not in temp_list:
                    temp_list.append(combo)
        combined.append(temp_list)
    return combined


# @jit(nopython=True)
def ideal(endings, ind):    # have this run seperately? speed up the process multiple times
    ideal_plays = [[], []]
    ties = endings[ind]
    for tie in ties:
        for side in tie:
            ideal_plays[tie.index(side)].append(side)
    return ideal_plays


# combine = ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 0)# for o
print(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 0))
# print(combine[0])
# print(combine[1])
# print('l', len(combine))
# print('Calculation Function Load Up: Complete \nNow Loading: Game Calculations')
# combine_x = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 0))
# print('Calculated: 33.33%')
# combine_o = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 1))
# print('Calculated: 66.66%')
combine_t = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 2))
print('Calculated: 99.99%')
# print(combine_t)


# @jit(nopython=True)
def mimicking_perfection(side, other):   # analyzing why the game was drawed?
    if side == x_blocks:
        other = other['o']
        side_key = 'x'
        other_ind = 1
        side_ind = 0
    else:
        other = other['x']
        side_key = 'o'
        other_ind = 0
        side_ind = 1
    mimicking = []
    for attempt in combine_t[other_ind]:
        similar_counter = 0
        for length in range(len(other)):
            if other[:length+1] == attempt[:length+1]:
                similar_counter = length+1
                print(attempt)
            else:
                break
        if similar_counter:
            package = [similar_counter, combine_t[side_ind][combine_t[other_ind].index(attempt)]]
            if mimicking:
                if mimicking[0][0] == package[0] and package not in mimicking:
                    mimicking.append(package)
                    mimicking.sort(reverse=True)
                elif mimicking[0][0] < package[0]:
                    mimicking = [package]
            else:
                mimicking.append(package)
    for mim in mimicking:
        similar_counter = 0
        for length in range(len(side[side_key])):
            if mim[1][:length+1] == side[side_key][:length+1]:
                similar_counter = length
            else:
                break
        mim[0] += similar_counter
    mimic = []
    if mimicking:
        base_line = sorted(mimicking)[-1][0]
        for mim in mimicking:
            if mim[0] == base_line:
                mimic.append(mim[1])
    return mimic


subject = mimicking_perfection(o_blocks, x_blocks)
print(subject)
if subject:
    print(True)
else:
    print(False)

# def reducing(loc=None, lvl2=None):
#     if not loc:
#         loc = input('path: ')
#     wb = xlrd.open_workbook(loc)
#     sheet = wb.sheet_by_index(0)
#     winning_x = []
#     winning_o = []
#     draw_counts = []
#     lists_of_plays = [winning_x, winning_o, draw_counts]
#     rows_to_do = [1, 2, 3]
#     for r in rows_to_do:
#         if not sheet.cell_value(2, r):
#             continue
#         for combo in range(int(sheet.cell_value(r, 4)) - 1):
#             com = sheet.cell_value(combo + 1, r)
#             if lvl2:
#                 com = com.replace('{', '').replace('}', '').replace("'x': ", '').replace("'o': ", '')
#             com = com[2:-2].replace("'", '').replace(', ', '').split('][')
#             comb = [[], []]
#             for block in com[0]:
#                 comb[0].append(block)
#             for block in com[1]:
#                 comb[1].append(block)
#             if comb not in lists_of_plays[rows_to_do.index(r)]:
#                 lists_of_plays[rows_to_do.index(r)].append(comb)
#     return lists_of_plays
#
#
# def combining(list1, list2, list3):
#     combined = []
#     lists = [list1, list2, list3]
#     indexes = (0, 1, 2)
#     for ind in indexes:
#         temp_list = []
#         for list_ in lists:
#             for combo in list_[ind]:
#                 if combo not in temp_list:
#                     temp_list.append(combo)
#         combined.append(temp_list)
#     return combined
#
#
# def ideal(endings):
#     # remember to take the [2] of combined lists because that is the draws
#     ideal_plays = [[], []]
#     ties = endings[-1]
#     for tie in ties:
#         for side in tie:
#             ideal_plays[tie.index(side)].append(side)
#     return ideal_plays
#
#
# def win_checker(placed, plc=None):
#     if placed == x_blocks:
#         placed = x_blocks['x']
#     else:
#         placed = o_blocks['o']
#     for win in wins:
#         counter = 0
#         for block in placed:
#             if block in win:
#                 counter += 1
#         if counter == 3:
#             return 'win'
#         if plc and counter == 2:
#             return win
#     return 'ongoing'
#
#
# wins = (['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'],
#         ['a', 'd', 'g'], ['b', 'e', 'h'], ['c', 'f', 'i'],
#         ['a', 'e', 'i'], ['c', 'e', 'g'])
# blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
# x_blocks = {'x': []}
# o_blocks = {'o': []}
# side = x_blocks
# possible_routes = []
# other = o_blocks
# combine = ideal(combining(reducing(location1), reducing(location2), reducing(location3)))
# if side == x_blocks:
#     side_ind = 0
# else:
#     side_ind = 1
# for opportunity in combine[side_ind]:
#     print(opportunity)
#     for needed in opportunity:
#         usable = True
#         if needed in side or needed in other:
#             usable = False
#             break
#     if usable:
#         possible_routes.append(opportunity)
# print(possible_routes)  # works fine, but i need to refresh every list before appending, making a list too long, have
# 2 seperate lists? just like before, lists with in list, but each list within has lists of combinations
# improve by stopping once found a working solution, but that is for later generations, but they would be faster anyways
# loc = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\dataset_gen2.0.xlsx'
# wb = xlrd.open_workbook(loc)
# sheet = wb.sheet_by_index(0)
# comb = sheet.cell_value(1, 1).replace("{'x'", '').replace("{'o'", '').replace(': ', '').replace("}", '')[1:-1].replace("'", '').split('], [')
# temp = sheet.cell_value(1, 1).replace("{'x'", '').replace("{'o'", '').replace(': ', '').replace("}", '')[2:-2].split('], [')
# print(temp)
# print(type(temp))

# x_ = {'x': []}
# o_ = {'o': []}
#
#
# def test(x, o):
#     if x is x_ and o is o_:
#         return True
#
#
# print(test(x_, x_))
# location1 = "C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\collection_1.xlsx"
# wb = xlrd.open_workbook(location1)
# sheet = wb.sheet_by_index(0)
# sheet.cell_value(0, 0)
# print(sheet.cell_value(1, 3))
# l1 = [0.2, 0.3333333333333333, 0.2727272727272727, 0.26666666666666666, 0.17857142857142858, 0.10344827586206896, 0.1044776119402985, 0.10666666666666667, 0.11392405063291139, 0.11363636363636363, 0.12087912087912088, 0.11214953271028037, 0.10236220472440945, 0.10606060606060606, 0.09090909090909091, 0.050473186119873815]
# l2 = [0.14285714285714285, 0.11764705882352941, 0.07317073170731707, 0.08, 0.09090909090909091, 0.1016949152542373, 0.06862745098039216, 0.07272727272727272, 0.072, 0.072992700729927, 0.07913669064748201, 0.07100591715976332, 0.06914893617021277, 0.07253886010362694, 0.06880733944954129, 0.07207207207207207]
# l3 = [0.3333333333333333, 0.18181818181818182, 0.1875, 0.18181818181818182, 0.14285714285714285, 0.13333333333333333, 0.1044776119402985, 0.11428571428571428, 0.1267605633802817, 0.07575757575757576, 0.08148148148148149, 0.0851063829787234, 0.08843537414965986, 0.08536585365853659, 0.03978779840848806, 0.03470715835140998]
#
# final_list = []
# for ind in range(0, 16):
#     final_list.append((l1[ind]+l2[ind]+l3[ind])/3)
# print(final_list)
# ws['E1'] = game_count/10
# ws['E2'] = lx
# ws['E3'] = lo
# ws['E4'] = ld
# from basic_tic_tac_toe_playing import block_choosing
# C:\Users\Dell\Desktop\Academic\Comp_Sci\SRP\data\collection_1.xlsx

# x = None
# if not x:
#     print('not')
# def reducing():
#     loc = input('path: ')
#     wb = xlrd.open_workbook(loc)
#     sheet = wb.sheet_by_index(0)
#     # sheet.cell_value(0, 0)
#     winning_x = []
#     winning_o = []
#     draw_counts = []
#     lists_of_plays = [winning_x, winning_o, draw_counts]
#     rows_to_do = [1, 2, 3]
#     for r in rows_to_do:
#         for combo in range(int(sheet.cell_value(r, 4))):
#             comb = sheet.cell_value(combo+1, 1)[1:-1].replace("'", '').split('], [')
#             for side in comb:
#                 if '[' in side:
#                     comb[comb.index(side)] = side.replace('[', '').split(', ')
#                 else:
#                     comb[comb.index(side)] = side.replace(']', '').split(', ')
#             if comb not in lists_of_plays[rows_to_do.index(r)]:
#                 lists_of_plays[rows_to_do.index(r)].append(comb)


# time = str(datetime.datetime.now()).replace(' ', '_').split('.')
# print(type(time[0]), time[0])
# string = 'h e, l, l o'
# string = string[1:-1]
# print(string)

# wins = [['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'],
#         ['a', 'd', 'g'], ['b', 'e', 'h'], ['c', 'f', 'i'],
#         ['a', 'e', 'i'], ['c', 'e', 'g']]
# tactics = []
# for win in wins:
#     for win_ind in range(wins.index(win) + 1, len(wins)):
#         temp_tac = []
#         tact_ava = False
#         for block in win:
#             if block in wins[win_ind]:
#                 tact_ava = True
#                 break
#         if tact_ava:
#             for block in win:
#                 # temp_tac = np.append(temp_tac, block)
#                 temp_tac.append(block)
#             for block in wins[win_ind]:
#                 if block not in temp_tac:
#                     temp_tac.append(block)
#                     # temp_tac = np.append(temp_tac, block)
#                 else:
#                     temp_tac.insert(0, [block])
#                     # temp_tac = np.insert(temp_tac, [0], [block])
#         if len(temp_tac):
#             tactics.append(temp_tac)
# # solution creating
# existing = []
# li = []
# for tactic in tactics:
#     li.append(tactic[0])
#     if tactic[0] not in existing:
#         existing.append(tactic[0])
# tactics_z = []
# for element in existing:
#     goin_in2_l = [element]
#     for tactic in tactics:
#         if tactic[0] == element:
#             for block in tactic[1:]:
#                 if block not in goin_in2_l:
#                     goin_in2_l.append(block)
#     tactics_z.append(goin_in2_l)
# options = {}
# for option in existing:
#     for item in tactics_z:
#         if item[0] == option:
#             options.update({f'{option}': item})
# print(options)
