# only checks for immediate dangers
# does not defend against forks
# do try and except for all possible routes
# take routes from the data collected
# 10001 ratios
import random as rand
from typing import List
import numpy
import matplotlib.pyplot as plt
from openpyxl import Workbook
import functions

wb = Workbook()
ws = wb.active

wins: List[List[str]] = [['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'],
                         ['a', 'd', 'g'], ['b', 'e', 'h'], ['c', 'f', 'i'],
                         ['a', 'e', 'i'], ['c', 'e', 'g']]
blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
x_blocks = []
o_blocks = []


def win_checker(placed, plc=None):
    for win in wins:
        counter = 0
        for block in placed:
            if block in win:
                counter += 1
        if counter == 3:
            return 'win'
        if plc and counter == 2:
            return win  # blocking or finishing
    return 'ongoing'


def block_choosing(side, other):    # i will need to find a way to import
    for block in x_blocks:  # checking for dupes
        if block in blocks:
            blocks.pop(blocks.index(block))
    for block in o_blocks:  # checking for dupes
        if block in blocks:
            blocks.pop(blocks.index(block))

    if type(win_checker(side, 'plc')) == list:  # if next to win
        for block in win_checker(side, 'plc'):
            if block in blocks:
                side.append(blocks.pop(blocks.index(block)))
                return
    if type(win_checker(other, 'opp')) == list:  # if there are no wins and the opp is one away
        for block in win_checker(other, 'opp'):
            if block in blocks:
                side.append(blocks.pop(blocks.index(block)))
                return
    # be careful with testing, if the blocks are "one away", this part would not be reached

    possible_routes = []
    for win in wins:
        go_for = True
        for needed in win:
            if needed in side or needed in other:
                go_for = False
        if go_for:
            possible_routes.append(win)

    if possible_routes:
        side.append(blocks.pop(blocks.index(rand.choice(rand.choice(possible_routes)))))
        return
    elif not possible_routes and blocks:
        side.append(blocks.pop(blocks.index(rand.choice(blocks))))
        return
    else:
        # print('tie')
        return 'tie'


# variables
x_w = 0
o_w = 0
tie = 0
every_10 = []
succeeded_x = []
succeeded_o = []
draw = []
game_count = int(input('how many trials of simulation? above 1000: '))

for game in range(game_count):  # in x many games, the win to l changed as such
    blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
    x_blocks = []
    o_blocks = []
    ongoing = True
    while ongoing:
        # block_choosing(x_blocks, o_blocks)
        x_blocks.append(blocks.pop(blocks.index(rand.choice(blocks))))
        # x_blocks.append(blocks.pop(blocks.index(input('x: '))))
        # print(f'x: {x_blocks} || o: {o_blocks}')
        # print(f'x: {x_blocks} || o: {o_blocks}')
        if win_checker(x_blocks) == 'win':
            # print(f'x wins, {x_blocks}')
            succeeded_x.append([x_blocks, o_blocks])
            x_w += 1
            ongoing = False
            break

        if not blocks and win_checker(x_blocks) != 'win' and win_checker(o_blocks) != 'win':
            # print(f'draw: {x_blocks, o_blocks}')
            draw.append([x_blocks, o_blocks])
            tie += 1
            ongoing = False
            break

        # block_choosing(o_blocks, x_blocks)
        o_blocks.append(blocks.pop(blocks.index(rand.choice(blocks))))
        if win_checker(o_blocks) == 'win':
            # print(f'o wins, {o_blocks}')
            succeeded_o.append([x_blocks, o_blocks])
            o_w += 1
            ongoing = False

        if not blocks and win_checker(x_blocks) != 'win' and win_checker(o_blocks) != 'win':
            # print(f'draw: {x_blocks, o_blocks}')
            draw.append([x_blocks, o_blocks])
            tie += 1
            ongoing = False

    if not (game + 1) % 10:
        every_10.append([x_w, o_w, tie])
        if not (game+1) % 1000:
            print(f'{(game+1)/1000}%')

headers = ['x:o:t', 'x_wins', 'o_wins', 'draws']
rows = ['A', 'B', "C", 'D']
row_count = 0
variables_to_write = [every_10, succeeded_x, succeeded_o,  draw]
lx = len(succeeded_x)
lo = len(succeeded_o)
ld = len(draw)
for row in range(len(headers)):
    ws[f'{rows[row]}{1}'] = headers[row]
for variable in variables_to_write:
    line_count = 2
    for item in variable:
        ws[f'{rows[row_count]}{line_count}'] = str(item)
        line_count += 1
    row_count += 1
ws['E1'] = game_count/10
ws['E2'] = lx
ws['E3'] = lo
ws['E4'] = ld
wb.save('bf2.xlsx')
