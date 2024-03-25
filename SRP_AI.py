# SRP AI
import random as rand
from openpyxl import Workbook
import xlrd
from numba import jit, cuda

# SRP AI
# Using genetic algorithm and other ways of calculation, this AI will find the best solutions to all questions asked
# The idea of this AI is that it gives a marking variable to all it's variables, then finds the best one overall
# the main purpose of this AI is to grow, everything else is but a way of measurement, eventually becoming an AGI
# use a huge tree of dictionary and lists of words to create a neural network of sort

# scoring system? marking blocks and also which ones are best for which steps
# fix the small numbering issues
# what if o has no wins
# make dictionary of words with syns as definition, and each syn its own
# Data refinement: save the ideal and combined and reduced into a file, and keep expanding on that file

# location1 = "C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\dataset_gen4.0.xlsx"
# location2 = "C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\dataset_gen4.1.xlsx"
# location3 = "C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\dataset_gen4.2.xlsx"

location1 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\blindtest0.xlsx'
location2 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\blindtest1.xlsx'
location3 = 'C:\\Users\\Dell\\Desktop\\Academic\\Comp_Sci\\SRP\\data\\blindtest2.xlsx'

wb1 = Workbook()
ws = wb1.active

# should i put this as a resetting function?
wins = (['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'],  # requirements, not possibilities
        ['a', 'd', 'g'], ['b', 'e', 'h'], ['c', 'f', 'i'],
        ['a', 'e', 'i'], ['c', 'e', 'g'])
blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
x_blocks = {'x': []}
o_blocks = {'o': []}


# @jit(target=cuda)
def reducing(loc=None, lvl2=None):
    if not loc:
        loc = input('path: ')
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    winning_x = []
    winning_o = []
    draw_counts = []
    lists_of_plays = [winning_x, winning_o, draw_counts]
    rows_to_do = [1, 2, 3]
    for r in rows_to_do:
        if not sheet.cell_value(2, r):
            continue
        for combo in range(int(sheet.cell_value(r, 4)) - 1):
            com = sheet.cell_value(combo + 1, r)
            if lvl2:
                com = com.replace('{', '').replace('}', '').replace("'x': ", '').replace("'o': ", '')
            com = com[2:-2].replace("'", '').replace(', ', '').split('][')
            # comb = [[list(com[0])], [list(com[1])]]
            comb = [[], []]
            for block in com[0]:
                comb[0].append(block)
            for block in com[1]:
                comb[1].append(block)
            if comb not in lists_of_plays[rows_to_do.index(r)]:  # do a for loop checking if there are ones dif order
                lists_of_plays[rows_to_do.index(r)].append(comb)
    return lists_of_plays


# @jit(target=cuda)
def combining(list1, list2, list3):
    combined = []
    lists = [list1, list2, list3]
    indexes = (0, 1, 2)
    for ind in indexes:
        temp_list = []
        for list_ in lists:
            for combo in list_[ind]:
                if combo not in temp_list:
                    temp_list.append(combo)
        combined.append(temp_list)
    return combined


# @jit(target=cuda)
def ideal(endings, ind):
    ideal_plays = [[], []]
    ties = endings[ind]
    for tie in ties:
        for side in tie:
            ideal_plays[tie.index(side)].append(side)
    return ideal_plays


# make these guys into some kind of knowledge excel file?
print('Calculation Function Load Up: Complete \nNow Loading: Game Calculations')
combine_x = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 0))
print('Calculated: 33.33%')
combine_o = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 1))
print('Calculated: 66.66%')
combine_t = tuple(ideal(combining(reducing(location1), reducing(location2), reducing(location3)), 2))
print('Calculated: 99.99%')


def win_checker(placed, plc=None):
    if placed == x_blocks:
        placed = x_blocks['x']
    else:
        placed = o_blocks['o']
    do_ables = []
    for win in wins:
        counter = 0
        for block in placed:
            if block in win:
                counter += 1
        if counter == 3:
            return 'win'
        if plc and counter == 2:
            do_ables.append(win)
    if do_ables:
        return do_ables
    return 'ongoing'


# @jit(target=cuda)
def mimicking_perfection(side, other):   # if side has already taken a spot, increase points
    if side == x_blocks:
        other = other['o']
        side_key = 'x'
        other_ind = 1
        side_ind = 0
    else:
        other = other['x']
        side_key = 'o'
        other_ind = 0
        side_ind = 1
    mimicking = []
    for attempt in combine_t[other_ind]:
        similar_counter = 0
        for length in range(len(other)):
            if other[:length+1] == attempt[:length+1]:
                similar_counter = length+1
            else:
                break
        if similar_counter:
            package = [similar_counter, combine_t[side_ind][combine_t[other_ind].index(attempt)]]
            if mimicking:
                if mimicking[0][0] == package[0] and package not in mimicking:
                    mimicking.append(package)
                    mimicking.sort(reverse=True)
                elif mimicking[0][0] < package[0]:
                    mimicking = [package]
            else:
                mimicking.append(package)
    for mim in mimicking:
        similar_counter = 0
        for length in range(len(side[side_key])):
            if mim[1][:length+1] == side[side_key][:length+1]:
                similar_counter = length
            else:
                break
        mim[0] += similar_counter
    mimic = []
    if mimicking:
        base_line = sorted(mimicking)[-1][0]
        for mim in mimicking:
            if mim[0] == base_line:
                mimic.append(mim[1])
    return mimic


# @jit(target=cuda)
def block_choosing(side, learned=None):  # should just randomly pick the first x move...., and start with o b_cing
    # side determining
    if side == x_blocks:
        side_ind = 0
        side = x_blocks
        side_key = 'x'
        other = o_blocks
    else:
        side_ind = 1
        side = o_blocks
        side_key = 'o'
        other = x_blocks

    # checking for dupes
    for block in x_blocks['x']:
        if block in blocks:
            blocks.pop(blocks.index(block))
    for block in o_blocks['o']:
        if block in blocks:
            blocks.pop(blocks.index(block))

    # hierarchy pt 1
    if type(win_checker(side, 'plc')) == list:  # if next to win
        for doable in win_checker(side, 'plc'):
            for block in doable:
                if block in blocks:
                    # print('checkmate')
                    side[side_key].append(blocks.pop(blocks.index(block)))
                    return
    if type(win_checker(other, 'opp')) == list:  # if there are no wins and the opp is one away
        # print('defended')
        for doable in win_checker(other, 'opp'):
            for block in doable:
                if block in blocks:
                    side[side_key].append(blocks.pop(blocks.index(block)))
                    return

    # hierarchy pt 2
    possible_routes = []
    if learned:     # is this really necessary?
        if not side_ind:  # side_ind is 0
            possible_routes = mimicking_perfection(x_blocks, o_blocks)
        else:
            possible_routes = mimicking_perfection(o_blocks, x_blocks)
    if not possible_routes:
        for opportunity in wins:  # random sampling to make it faster?
            top_priority = False
            usable = 0
            for needed in opportunity:
                if needed in side:
                    top_priority = True
                if needed in other:  # if trying to tie, wouldn't having needed in other also work?
                    usable += 1
            if usable <= 2:
                if top_priority:
                    possible_routes = [opportunity]
                    break
                possible_routes.append(opportunity)
    if possible_routes:     # need random
        rate = []
        for possible in possible_routes:
            for block in possible:
                rate.append(block)
        chosen_from = []
        for block in set(rate):
            chosen_from.append((rate.count(block), block))
        base_line = sorted(chosen_from)[-1][0]
        similar_scores = []
        for block in chosen_from:
            if block[1] in blocks and block[0] == base_line:
                similar_scores.append(block[1])
        if similar_scores:
            # print(similar_scores)
            side[side_key].append(blocks.pop(blocks.index(rand.choice(similar_scores))))
            return
    if blocks:
        # print('no routes')
        side[side_key].append(blocks.pop(blocks.index(rand.choice(blocks))))
        return
    return 'tie'


print('Game Function Load Up: Complete')


def saving_data(header, rows_, variables_to_write, gen=None):
    row_count = 0
    for row in range(len(header)):
        ws[f'{rows_[row]}{1}'] = header[row]
    for var in variables_to_write:
        line_count = 2
        for variable in var:
            if variable and type(variable[0]) == dict:
                ws[f'{rows_[row_count]}{line_count}'] = str([variable[0]['x'], variable[1]['o']])
                line_count += 1
            else:
                ws[f'{rows_[row_count]}{line_count}'] = str(variable)
                line_count += 1
        row_count += 1
        if (line_count - 2) > 0:
            ws[f'E{row_count}'] = (line_count - 2)
        else:
            ws[f'E{row_count}'] = 0
    ws['E1'] = int(game_count/10)+1
    wb1.save(f'dataset_gen{gen}.xlsx')


x_w = 0
o_w = 0
draw = 0
every_10 = []
succeeded_x = []
succeeded_o = []
draws = []
game_count = int(input('Game counts (large number): '))
for run in range(game_count):
    blocks = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
    x_blocks = {'x': []}
    o_blocks = {'o': []}
    x_blocks['x'].append(blocks.pop(blocks.index(input(blocks))))
    # x_blocks['x'].append(blocks.pop(blocks.index(rand.choice(blocks))))
    ongoing = True
    while ongoing:
        block_choosing(o_blocks, 'learned')
        # block_choosing(x_blocks)
        # print(x_blocks)
        print(o_blocks)
        if win_checker(o_blocks) == 'win':
            # print('x wins', x_blocks, o_blocks)
            succeeded_o.append([x_blocks, o_blocks])
            o_w += 1
            ongoing = False
            break
        x_blocks['x'].append(blocks.pop(blocks.index(input(blocks))))
        # block_choosing(x_blocks, 'learned')
        # block_choosing(o_blocks)
        # print(o_blocks)
        if win_checker(x_blocks) == 'win':
            # print('o, wins', x_blocks, o_blocks)
            succeeded_x.append([x_blocks, o_blocks])
            x_w += 1
            ongoing = False
            break
        if not blocks and win_checker(x_blocks) != 'win' and win_checker(o_blocks) != 'win':
            # print(f'draw: {x_blocks, o_blocks}')
            draws.append([x_blocks, o_blocks])
            draw += 1
            ongoing = False
    if not (run + 1) % 10:
        every_10.append([x_w, o_w, draw])
        if not (run+1) % 1000:
            print(f'{(run+1)*100/game_count}%')
if not every_10:
    every_10 = [[x_w, o_w, draw]]

saving_data(['x:o:t', 'x', 'o', 't'], ['A', 'B', "C", 'D'],
            [every_10, succeeded_x, succeeded_o, draws], gen='3.2')
